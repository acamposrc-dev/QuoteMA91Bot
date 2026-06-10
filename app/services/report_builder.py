"""Genera el Excel de resultados: una fila por opcion, agrupado por item."""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.config import settings
from app.models import QuoteRequest

TIER_NAMES = {"ve": "Venezuela", "us": "EEUU", "cn": "China", "global": "Internacional"}
LEVEL_NAMES = {"exact": "Exacto", "equivalent": "Equivalente", "partial": "Parcial"}

HEADERS = ["Item", "Cant.", "Opcion", "Producto encontrado", "Vendedor", "Origen",
           "Precio", "Moneda", "Precio USD", "Total USD", "Equivalencia", "Justificacion", "Link"]


def build_report(request: QuoteRequest) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotizacion"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    row = 2
    for item in request.items:
        selected = sorted([o for o in item.options if o.rank], key=lambda o: o.rank)
        if not selected:
            ws.cell(row=row, column=1, value=item.name)
            ws.cell(row=row, column=2, value=item.quantity)
            note = ws.cell(row=row, column=4,
                           value="SIN OPCIONES: ninguna oferta encontrada cumple las especificaciones")
            note.font = Font(italic=True, color="9C0006")
            row += 1
            continue
        if len(selected) < settings.options_per_item:
            ws.cell(row=row, column=4,
                    value=f"NOTA: solo {len(selected)} opcion(es) cumplen las especificaciones"
                    ).font = Font(italic=True, color="9C6500")
            ws.cell(row=row, column=1, value=item.name)
            ws.cell(row=row, column=2, value=item.quantity)
            row += 1
        for opt in selected:
            ws.cell(row=row, column=1, value=item.name)
            ws.cell(row=row, column=2, value=item.quantity)
            ws.cell(row=row, column=3, value=opt.rank)
            ws.cell(row=row, column=4, value=opt.title)
            ws.cell(row=row, column=5, value=opt.seller)
            ws.cell(row=row, column=6, value=TIER_NAMES.get(opt.tier, opt.tier))
            ws.cell(row=row, column=7, value=opt.price_amount)
            ws.cell(row=row, column=8, value=opt.price_currency)
            ws.cell(row=row, column=9, value=opt.price_usd)
            ws.cell(row=row, column=10, value=round(opt.price_usd * item.quantity, 2))
            ws.cell(row=row, column=11, value=LEVEL_NAMES.get(opt.equivalence_level, opt.equivalence_level))
            ws.cell(row=row, column=12, value=opt.equivalence_reason)
            link = ws.cell(row=row, column=13, value=opt.url)
            link.hyperlink = opt.url
            link.font = Font(color="0563C1", underline="single")
            row += 1

    widths = [30, 6, 7, 50, 22, 14, 12, 8, 12, 12, 14, 45, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(vertical="top", wrap_text=(c.column in (4, 12)))

    os.makedirs(os.path.join(settings.storage_dir, "reports"), exist_ok=True)
    path = os.path.join(settings.storage_dir, "reports", f"cotizacion_{request.id}.xlsx")
    wb.save(path)
    return path
