"""Parser deterministico de la plantilla oficial Excel.

Plantilla esperada (hoja 1, fila 1 = encabezados, case-insensitive):
    item | descripcion | cantidad | marca_preferida | specs
La columna specs admite pares "clave: valor" separados por ";".
"""

import io

from openpyxl import load_workbook

from app.models import ParseMethod
from app.services.parsing.base import DocumentParser, ParsedItem, ParseResult, UnrecognizedFormat

REQUIRED = {"item", "descripcion", "cantidad"}
OPTIONAL = {"marca_preferida", "specs"}


def parse_specs(raw: str | None) -> dict[str, str]:
    if not raw:
        return {}
    specs: dict[str, str] = {}
    for chunk in str(raw).split(";"):
        if ":" in chunk:
            k, v = chunk.split(":", 1)
            specs[k.strip().lower()] = v.strip()
        elif chunk.strip():
            specs[chunk.strip().lower()] = "true"
    return specs


class TemplateXlsxParser(DocumentParser):
    method = ParseMethod.TEMPLATE_XLSX.value

    def can_handle(self, filename: str, content: bytes) -> bool:
        return filename.lower().endswith((".xlsx", ".xlsm"))

    def parse(self, filename: str, content: bytes) -> ParseResult:
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise UnrecognizedFormat(f"No es un xlsx valido: {exc}") from exc

        ws = wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            raise UnrecognizedFormat("Hoja vacia")

        cols = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}
        if not REQUIRED.issubset(cols):
            raise UnrecognizedFormat(
                f"Encabezados no matchean la plantilla. Encontrados: {sorted(cols)}"
            )

        items: list[ParsedItem] = []
        for row in rows:
            name = row[cols["item"]]
            if name is None or not str(name).strip():
                continue
            qty_raw = row[cols["cantidad"]]
            try:
                qty = max(1, int(float(qty_raw))) if qty_raw is not None else 1
            except (TypeError, ValueError):
                raise UnrecognizedFormat(f"Cantidad invalida en fila: {row}")
            items.append(
                ParsedItem(
                    name=str(name).strip(),
                    description=str(row[cols["descripcion"]] or "").strip() or None,
                    quantity=qty,
                    preferred_brand=(
                        str(row[cols["marca_preferida"]] or "").strip() or None
                        if "marca_preferida" in cols
                        else None
                    ),
                    specs=parse_specs(row[cols["specs"]] if "specs" in cols else None),
                )
            )
        if not items:
            raise UnrecognizedFormat("Plantilla reconocida pero sin items")
        return ParseResult(items=items, method=self.method)
